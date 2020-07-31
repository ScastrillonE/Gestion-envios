from django.http import HttpResponse
from django.shortcuts import render
from datetime import timedelta
import datetime
from django.utils.dateparse import parse_date
from openpyxl import Workbook

from .models import Shipping


def reportes(request, f1, f2):
    if request.method == 'GET':
        f2 = parse_date(f2)
        f2 = f2 + timedelta(days=1)
        print(f1)
        print(f2)
        envios = Shipping.objects.filter(send_date__gte=f1, send_date__lt=f2,status = 'Enviado'
                                                                                      '')
        f2 = f2 - timedelta(days=1)

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Reporte creado: {}'.format(datetime.datetime.now())
        ws.merge_cells('A1:B1')

        ws['A2'] = 'Cliente'
        ws['B2'] = 'Tipo de documento'
        ws['C2'] = 'Documento'
        ws['D2'] = 'Direccion'
        ws['E2'] = 'Telefono'
        ws['F2'] = 'Celular'
        ws['G2'] = 'Empresa de envio'
        ws['H2'] = 'Fecha de envio'
        ws['I2'] = 'Estado'

        cont = 3

        for envio in envios:
            ws.cell(row=cont, column=1).value = envio.get_customer
            ws.cell(row=cont, column=2).value = envio.c_Tdocument
            ws.cell(row=cont, column=3).value = envio.c_document
            ws.cell(row=cont, column=4).value = envio.c_address
            ws.cell(row=cont, column=5).value = envio.c_phone
            ws.cell(row=cont, column=6).value = envio.c_cel
            ws.cell(row=cont, column=7).value = envio.get_company
            ws.cell(row=cont, column=8).value = envio.send_date
            ws.cell(row=cont, column=9).value = envio.status
            cont += 1

        nombre = 'Reporte envios.xlsx'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {}".format(nombre)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
